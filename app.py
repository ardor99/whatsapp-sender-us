import os
import threading
import openpyxl
import re
import pyperclip
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, send_from_directory
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Main Page
@app.route('/')
def index():
    return render_template('index.html')

# WhatsApp Sender
@app.route('/whatsapp_sender', methods=['GET', 'POST'])
def whatsapp_sender():
    if request.method == 'POST':
        file = request.files['file']
        language = request.form['language']
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        whatsapp_thread = threading.Thread(target=start_whatsapp_automation, args=(file_path, language))
        whatsapp_thread.start()

        return redirect(url_for('status', tool='whatsapp_sender'))

    return render_template('whatsapp_sender.html')

def start_whatsapp_automation(filepath, language):
    # Ensure WebDriver uses the VNC display
    os.environ["DISPLAY"] = ":1"  # Or the display number your VNC server is using

    # Set Chrome options
    chrome_options = Options()
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    # Remove headless mode to allow browser window to be visible
    # chrome_options.add_argument("--headless")  # Ensure this is commented out or removed

    # Start Chrome WebDriver
    service = Service("/usr/local/bin/chromedriver")
    driver = webdriver.Chrome(service=service, options=chrome_options)

    # Open WhatsApp Web
    driver.get("https://web.whatsapp.com/")

    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    wb_sent = openpyxl.Workbook()
    sheet_sent = wb_sent.active
    sheet_sent.append(["Phone Number", "Message", "Status"])

    wb_unsent = openpyxl.Workbook()
    sheet_unsent = wb_unsent.active
    for row in sheet.iter_rows(values_only=True):
        sheet_unsent.append(row)

    unsent_row = 2

    for row in sheet.iter_rows(min_row=2, values_only=True):
        phone_number, message = str(row[0]), str(row[1])

        try:
            driver.get(f"https://web.whatsapp.com/send?phone={phone_number}")
            print(f"Attempting to send message to {phone_number}...")

            message_box_xpath = {
                "en": "//div[@aria-label='Type a message']",
                "ar": "//div[@aria-label='اكتب رسالة']"
            }[language]

            # Wait for the message box to be available
            message_box = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, message_box_xpath)))

            # Clean the message
            cleaned_message = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\xff]', '', message)
            pyperclip.copy(cleaned_message)

            # Paste and send the message
            message_box.send_keys(Keys.CONTROL, 'v')
            message_box.send_keys(Keys.ENTER)

            # Check message status
            icons = [
                ("//span[@data-icon='msg-check']", "Sent"),
                ("//span[@data-icon='msg-dblcheck']", "Delivered"),
                ("//span[@data-icon='msg-dblcheck-ack']", "Read")
            ]

            status = "Unknown"
            for icon, state in icons:
                try:
                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, icon)))
                    status = state
                    break
                except:
                    continue

            sheet_sent.append([phone_number, message, status])
            sheet_unsent.delete_rows(unsent_row)

            print(f"Message to {phone_number} {status}.")

        except Exception as e:
            print(f"Error for {phone_number}: {str(e)}")
            unsent_row += 1

        time.sleep(2)  # Pause between messages

    driver.quit()

    # Save the sent and unsent messages
    wb_sent.save(os.path.join(app.config['UPLOAD_FOLDER'], "sent_messages.xlsx"))
    wb_unsent.save(os.path.join(app.config['UPLOAD_FOLDER'], "unsent_messages.xlsx"))

# XLS to CSV Converter
@app.route('/xls_to_csv', methods=['GET', 'POST'])
def xls_to_csv():
    if request.method == 'POST':
        file = request.files['file']
        option = request.form['option']
        file_name = request.form['file_name']
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        output_file = process_xls_to_csv(file_path, file_name, option)

        return redirect(url_for('download_file', filename=output_file))

    return render_template('xls_to_csv.html')

def process_xls_to_csv(source_path, file_name, option):
    df = pd.read_excel(source_path, engine='openpyxl')
    df = df.dropna(subset=['الخليوي'])

    replacements = {
        'ف1': 'فئة اولى',
        'ف2': 'فئة ثانية',
        'ف3': 'فئة ثالثة',
        '1': 'اول',
        '2': 'ثاني',
        '3': 'ثالث',
        '4': 'رابع',
        '5': 'خامس',
        '6': 'سادس',
        '7': 'سابع',
        '8': 'ثامن',
        '9': 'تاسع',
        '5 ب': 'خامس ب',
        '6 ب': 'سادس ب',
        '7 ب': 'سابع ب',
        '8 ب': 'ثامن ب',
        '9 ب': 'تاسع ب'
    }
    df['الصف'] = df['الصف'].replace(replacements)

    column_2 = "Mobile Phone"
    column_1 = "First Name"
    df_empty = pd.DataFrame({column_1: [], column_2: []})

    df_empty['Mobile Phone'] = "+" + df['الخليوي'].apply(lambda x: str(int(x)))

    if option == "الرقم المالي":
        df_empty['First Name'] = df['الصف'].astype(str) + "_" + df['الرقم المالي'].astype(str) + "_ " + df['الاسم'].astype(str) + " " + df['اللقب'].astype(str)
    else:
        df_empty['First Name'] = df['الصف'].astype(str) + "_" + df['الشعبة'].astype(str) + "_ " + df['الاسم'].astype(str) + " " + df['اللقب'].astype(str)

    df_empty = df_empty.groupby('Mobile Phone').apply(lambda group: group.iloc[0])
    df_empty.drop_duplicates(subset='Mobile Phone', keep='first', inplace=True)

    output_file = f"{file_name}.csv"
    destination_path = os.path.join(app.config['UPLOAD_FOLDER'], output_file)
    df_empty.to_csv(destination_path, index=False, encoding='utf-8-sig')
    
    return output_file

@app.route('/download/<filename>')
def download_file(filename):
    directory = app.config['UPLOAD_FOLDER']
    file_path = os.path.join(directory, filename)
    if os.path.exists(file_path):
        response = send_from_directory(directory, filename, as_attachment=True)
        os.remove(file_path)
        return response
    return "File not found", 404

# SMS Processor
@app.route('/sms_processor', methods=['GET', 'POST'])
def sms_processor():
    if request.method == 'POST':
        file = request.files['file']
        file_name = request.form['file_name']
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)

        output_file = process_sms(file_path, file_name)

        return redirect(url_for('download_file', filename=output_file))

    return render_template('sms_processor.html')

def process_sms(source_path, file_name):
    df = pd.read_excel(source_path, dtype={0: str})
    valid_df, invalid_df = process_phone_numbers(df)

    max_chunk_length = 30
    split_columns = ['Message_Part_1', 'Message_Part_2', 'Message_Part_3', 'Message_Part_4']
    output_df = pd.DataFrame()
    output_df['PhoneNumber'] = valid_df.iloc[:, 0]

    for index, message in enumerate(valid_df.iloc[:, 1]):
        chunks = split_message(str(message), max_chunk_length)
        chunks += ["  "] * (4 - len(chunks))
        for i in range(4):
            output_df.loc[index, split_columns[i]] = chunks[i] if i < len(chunks) else "  "
    final_valid_df, final_invalid_df = validate_data(output_df)
    all_invalid_df = pd.concat([invalid_df, final_invalid_df])
    final_valid_df = fill_empty_with_spaces(final_valid_df)

    output_file = f"{file_name}.xlsx"
    destination_path = os.path.join(app.config['UPLOAD_FOLDER'], output_file)
    final_valid_df.to_excel(destination_path, index=False, header=False, engine='openpyxl')

    return output_file

def process_phone_numbers(df):
    valid_rows = df[df.iloc[:, 0].str.startswith('963')]
    invalid_rows = df[~df.iloc[:, 0].str.startswith('963')]
    valid_rows.iloc[:, 0] = valid_rows.iloc[:, 0].str.replace('963', '0', 1)
    return valid_rows, invalid_rows

def split_message(message, max_length):
    words = message.split()
    chunks = []
    current_chunk = words[0]

    for word in words[1:]:
        if len(current_chunk) + len(word) + 1 <= max_length:
            current_chunk += " " + word
        else:
            chunks.append(current_chunk)
            current_chunk = word
    chunks.append(current_chunk)

    return chunks

def validate_data(df):
    valid_rows = df[df['PhoneNumber'].str.len() == 10]
    invalid_rows = df[df['PhoneNumber'].str.len() != 10]

    for index, row in valid_rows.iterrows():
        if any(len(str(cell)) > 30 for cell in row):
            invalid_rows = pd.concat([invalid_rows, pd.DataFrame([row], columns=row.index)], ignore_index=True)
            valid_rows = valid_rows.drop(index)

    return valid_rows, invalid_rows

def fill_empty_with_spaces(df):
    columns_to_check = [1, 2, 3, 4]
    for col in columns_to_check:
        if col < len(df.columns):
            df.iloc[:, col] = df.iloc[:, col].apply(lambda x: " " if x == '  ' else x)
    return df

# Status Page
@app.route('/status/<tool>')
def status(tool):
    return f"Processing {tool}, please check back later."

if __name__ == "__main__":
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    app.run(host='0.0.0.0', port=8080)
